import openpyxl
import requests  
import urllib3
import json
import time
import re
from datetime import datetime
from typing import Dict, List, Tuple
import statistics
import os
import pandas as pd
import queue
import threading
from openpyxl import Workbook, load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class EnhancedMultiAIPlanner:
    
    def __init__(self, event_queue=None, tool_excel_path=None, param_excel_path=None):
        """Initialize planner with optional tool paths"""
        self.api_keys = {
            "gemini": "AIzaSyBkAlPjP5vod1OfnVBZLuM3ZEZCjjnh77A",
            "openai": "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",
            "together": "bf9c17f509b1d83357c71c949a6a4cf58eb1156822f8d871e33a48afe4abed66",
            "groq": "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",
            "replicate": "r8_MyWMqWXCAvv2lCQMnVmT93mSwcyxMw911RwKV",
            "fireworks": "fw_3ZPncCHpEY5JcwNi6tuWRwFi",
            "cohere": "tC192FBsocqEdLB4So7aWwFTNrMc0kyFeFU22U1h",
            "deepseek": "sk-c6a2fc59f57e414fa7b12050500c0b7b",
            "openrouter": "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",
            "writer": "VhFZdA7xO958vn9GGjUXQpG0CrF1054x",
            "huggingface": "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",
            "pawn": "pk-oCcKygMBXZSfkXDGvSaafmejVpmhgkAYMpZpUPfgpEFczFWS"
        }
        
        # Only API-based AI agents
        self.ai_agents = ["gemini", "openai", "together", "groq", "replicate", "fireworks", 
                         "cohere", "deepseek", "openrouter", "writer", "huggingface", "pawn"]
        
        # Store all plans and scores
        self.ai_plans = {}
        self.ai_scores = {}
        self.improvement_chain_results = {}
        
        # ADD: Tool management paths
        self.tool_excel_path = tool_excel_path or r"C:\Users\LENOVO\Desktop\tool_name.xlsx"
        self.param_excel_path = param_excel_path or r"C:\Users\LENOVO\Desktop\parameter.xlsx"
        
        # ADD these new tracking dictionaries
        self.fraud_detection_results = {}  # Track fraud checks per mini-plan
        self.tech_integration_suggestions = {}  # Track tech suggestions per mini-plan
        self.node_generation_results = {}  # Track generated nodes
        
        # Track complete process
        self.process_log = []
        self.tracking_file = r"C:\Users\LENOVO\Desktop\RE.txt"
        self.parameters_file = r"C:\Users\LENOVO\Desktop\para.xlsx"
        
        # Store input steps and aim
        self.initial_steps = []
        self.aim = ""
        self.event_queue = event_queue

        # ADD these two lines after existing initialization
        self.discovered_features = set()
        self.feature_history = []
        
        # Competition analysis storage
        self.aim_competition = None
        self.mini_plan_competitions = []
        self.summary_competition = None

    def parse_natural_language_to_structure(self, user_input: str) -> dict:
        """Convert natural language input to structured AIM + STEPS using Gemini"""
        try:
            self.log_process("🧠 Converting natural language to structured format...")
            
            prompt = f"""You are a goal-structuring AI.
Your job is to take any kind of instruction, idea, or statement — whether it's a simple sentence, chat message, 
project description, or casual note — and convert it into a clear, concise, and structured format 
with only two parts: AIM and STEPS.

### AIM:
- The main goal or objective expressed by the user.
- Keep it short, action-oriented, and outcome-focused.

### STEPS:
- A list of logical and actionable steps needed to achieve that AIM.
- Use simple, imperative sentences (Start with verbs like Create, Build, Connect, Analyze, etc.).
- Steps should be sequential and easy to follow.
- No need for sub-steps, resources, or technical details unless implied naturally.

### OUTPUT FORMAT (Plain text):
AIM: [clear one-line goal]  
STEPS:  
1. [step one]  
2. [step two]  
3. [step three]  
...

Now convert the following input into AIM and STEPS:

{user_input}"""

            response = self.call_ai_api("gemini", prompt)
            
            if response.startswith(("❌", "🚫")):
                return {"error": "Failed to parse natural language"}
            
            # Parse the response
            aim = ""
            steps = []
            
            lines = response.split('\n')
            current_section = None
            
            for line in lines:
                line = line.strip()
                
                if line.upper().startswith('AIM:'):
                    aim = line.replace('AIM:', '').replace('aim:', '').strip()
                    current_section = 'aim'
                elif line.upper().startswith('STEPS:'):
                    current_section = 'steps'
                elif current_section == 'steps' and line:
                    # Remove numbering and extract step
                    step = line.lstrip('0123456789.-) ').strip()
                    if step and len(step) > 3:
                        steps.append(step)
            
            if not aim or len(steps) == 0:
                return {"error": "Could not extract AIM and STEPS from response"}
            
            # Convert steps list to dictionary format
            steps_dict = {}
            for i, step in enumerate(steps, 1):
                steps_dict[f'step{i}'] = step
            
            self.log_process(f"✅ Parsed: AIM + {len(steps)} steps")
            
            return {
                "aim": aim,
                "steps": steps_dict,
                "raw_response": response
            }
            
        except Exception as e:
            self.log_process(f"❌ Error parsing natural language: {e}")
            return {"error": str(e)}

    def check_fraud_and_misinformation(self, mini_plan_text: str, ai_name: str, group_num: int, steps: List[str]) -> dict:
        """Check mini-plan for fraud, misinformation, or suspicious patterns"""
        try:
            prompt = f"""You are the Fraud & Misinformation Detection Node for an AI Planner.

INPUT:
{{
  "source_type": "text",
  "payload": {json.dumps(mini_plan_text)},
  "user_context": {{
    "ai_agent": "{ai_name}",
    "group": {group_num},
    "step_count": {len(steps)}
  }},
  "thresholds": {{
     "suspicion_score": 0.6,
     "fraud_amount": 10000
  }}
}}

TASKS (in order):
1. Classify whether the payload contains:
   - misinformation / false claims (content that is likely untrue, misleading, or unverifiable)
   - social-engineering / phishing attempt
   - transaction anomaly / likely money-fraud pattern
   - safe/benign

2. For any suspicious classification, produce:
   - suspicion_score (0.00–1.00)
   - reason_list (concise bullet points with evidence lines from payload)
   - risk_type: one of [misinformation, phishing, social_engineering, transaction_anomaly, data_tampering, other, safe]
   - suggested_action: one of [block_execute, require_human_review, low_alert_log_only, request_verification, throttle, escalate_authorities, allow]
   - confidence_level (low|medium|high)
   - recommended_verification_steps (list; e.g., ask for transaction receipt, verify KYC, cross-check with bank API, request 2FA)

3. If payload indicates an explicit instruction or plan to commit illegal acts (theft, laundering, violent crime), do NOT provide facilitation details. Instead set:
   - suggested_action = escalate_authorities
   - reason_list includes "potential criminal facilitation — block and escalate"

4. Always return a short one-line summary for logs and a JSON machine-readable output.

OUTPUT FORMAT (exact JSON, no extra text):
{{
  "suspicion_score": 0.0,
  "risk_type": "safe",
  "confidence_level": "high",
  "reason_list": ["reason 1"],
  "suggested_action": "allow",
  "recommended_verification_steps": [],
  "summary_line": "Short human readable summary."
}}

Respond ONLY with valid JSON, no markdown backticks or extra text."""

            response = self.call_ai_api("gemini", prompt)
            
            if response.startswith(("❌", "🚫")):
                return {"error": "Fraud detection API failed"}
            
            # Clean and parse JSON
            clean_response = response.strip()
            if clean_response.startswith("```json"):
                clean_response = clean_response.replace("```json", "").replace("```", "").strip()
            
            result = json.loads(clean_response)
            
            # Log detection result
            if result.get('suspicion_score', 0) > 0.6:
                self.log_process(f"⚠️ FRAUD ALERT - {ai_name} group {group_num}: {result.get('risk_type', 'unknown')} (score: {result.get('suspicion_score', 0)})")
            else:
                self.log_process(f"✅ Fraud Check - {ai_name} group {group_num}: Safe")
            
            # Send event
            if self.event_queue:
                self.event_queue.put({
                    'type': 'fraud_detection',
                    'data': {
                        'agent': ai_name,
                        'group': group_num,
                        'result': result
                    },
                    'timestamp': datetime.now().isoformat()
                })
            
            return result
            
        except Exception as e:
            self.log_process(f"❌ Fraud detection error: {e}")
            return {"error": str(e)}

    def generate_safety_node(self, fraud_result: dict, full_steps: List[str], flagged_step: str, ai_name: str, group_num: int) -> dict:
        """Generate a new safety node based on detected fraud/issue"""
        try:
            if fraud_result.get('suspicion_score', 0) < 0.6:
                return {"skipped": "No significant risk detected"}
            
            prompt = f"""You are the Node Generator AI of the AI Planner system.

INPUT:
{{
  "detected_issue": "{fraud_result.get('risk_type', 'unknown')}",
  "flagged_step": "{flagged_step}",
  "reason": {json.dumps(fraud_result.get('reason_list', []))},
  "suspicion_score": {fraud_result.get('suspicion_score', 0)},
  "recommended_action": "{fraud_result.get('suggested_action', 'require_review')}",
  "full_steps": {json.dumps(full_steps[:10])}
}}

TASKS:
1. Analyze the full list of steps and locate the flagged step in sequence.
2. Design a new specialized node that:
   - Prevents or verifies the risky behavior found in the flagged step.
   - Enhances the Planner's safety, learning, or compliance abilities.
3. Define the integration logic:
   - Ask: "After which step should this new node be attached?"
   - Ask: "Until which step should it operate or monitor?"
   - Ask: "Which existing nodes or layers should it interact with?"
4. Generate a structured node blueprint with full definition and integration instructions.

OUTPUT FORMAT (exact JSON, no extra text):
{{
  "node_name": "SafetyNodeName",
  "node_purpose": "Clear purpose",
  "input_requirements": ["requirement1"],
  "processing_logic": ["step1", "step2"],
  "output_format": "JSON schema description",
  "integration_steps": {{
     "attach_after_step": "step number or name",
     "end_monitor_at_step": "step number or name",
     "integration_sequence": ["sequence description"]
  }},
  "connections": ["layer1", "layer2"],
  "learning_rules": ["rule1"],
  "safety_guardrails": ["guardrail1"]
}}

Respond ONLY with valid JSON, no markdown backticks or extra text."""

            response = self.call_ai_api("gemini", prompt)
            
            if response.startswith(("❌", "🚫")):
                return {"error": "Node generation API failed"}
            
            clean_response = response.strip()
            if clean_response.startswith("```json"):
                clean_response = clean_response.replace("```json", "").replace("```", "").strip()
            
            result = json.loads(clean_response)
            
            self.log_process(f"🛡️ Generated Safety Node: {result.get('node_name', 'Unknown')} for {ai_name} group {group_num}")
            
            # Send event
            if self.event_queue:
                self.event_queue.put({
                    'type': 'node_generated',
                    'data': {
                        'agent': ai_name,
                        'group': group_num,
                        'node': result
                    },
                    'timestamp': datetime.now().isoformat()
                })
            
            return result
            
        except Exception as e:
            self.log_process(f"❌ Node generation error: {e}")
            return {"error": str(e)}

    def suggest_tech_integration(self, mini_plan_text: str, ai_name: str, group_num: int) -> dict:
        """Suggest modern tech tools and automation for the mini-plan"""
        try:
            prompt = f"""You are a Tech-Integration & Process-Simplification expert.

Input:
{mini_plan_text}

Goal:
Identify which modern technologies, AI tools, automation platforms, or APIs (e.g., ChatGPT, Claude, Scale AI, Midjourney, Photoshop, Zapier, n8n, Notion AI, etc.) can simplify, accelerate, or scale each step.

For each step:
1. Suggest the most relevant tools or tech stack.
2. Explain how each tool can improve speed, quality, or scalability.
3. Recommend any automation workflows or integrations (if applicable).
4. Mention if the task can be fully automated, semi-automated, or should remain manual.
5. (Optional) Give sample prompts or API flow ideas if AI can handle it.

Output format (exact JSON, no extra text):
{{
  "suggestions": [
    {{
      "step_number": 1,
      "original_step": "step description",
      "tools": ["tool1", "tool2"],
      "use_case": "how these tools help",
      "automation_level": "Auto|Semi|Manual",
      "bonus_tip": "optional extra advice"
    }}
  ],
  "summary": "Overall tech integration summary"
}}

Respond ONLY with valid JSON, no markdown backticks or extra text."""

            response = self.call_ai_api("gemini", prompt)
            
            if response.startswith(("❌", "🚫")):
                return {"error": "Tech integration API failed"}
            
            clean_response = response.strip()
            if clean_response.startswith("```json"):
                clean_response = clean_response.replace("```json", "").replace("```", "").strip()
            
            result = json.loads(clean_response)
            
            self.log_process(f"🔧 Tech Integration - {ai_name} group {group_num}: {len(result.get('suggestions', []))} suggestions")
            
            # Send event
            if self.event_queue:
                self.event_queue.put({
                    'type': 'tech_integration',
                    'data': {
                        'agent': ai_name,
                        'group': group_num,
                        'suggestions': result
                    },
                    'timestamp': datetime.now().isoformat()
                })
            
            return result
            
        except Exception as e:
            self.log_process(f"❌ Tech integration error: {e}")
            return {"error": str(e)}

    def extract_tools_from_software(self, software_name: str, excluded_tools: List[str] = None) -> dict:
        """Extract 10 tools at a time from software until all discovered - PROMPT WITH REAL VALUES"""
        if excluded_tools is None:
            excluded_tools = []
        
        # Build excluded tools string with actual values
        excluded_str = ', '.join(excluded_tools) if excluded_tools else 'None'
        
        # FIXED PROMPT - More specific and clear
        prompt = f"""You are a software expert analyzing {software_name}.

TASK: List exactly 10 unique tools, features, or commands available in {software_name}.

WHAT TO INCLUDE:
- Menu items (File, Edit, View, etc.)
- Toolbar tools and buttons
- Drawing/editing tools
- Panels and palettes
- Commands and shortcuts
- Settings and preferences
- Filters, effects, or transformations
- Export/import options
- Plugins or extensions

ALREADY LISTED (skip these): {excluded_str}

OUTPUT FORMAT (exactly like this):
[Tool 1]
Name: Exact tool name
Category/Location: Where to find it
Function: What it does
Sub-options: Any variants or settings
Example Usage: Brief example

[Tool 2]
Name: ...

...continue for 10 tools total...

IMPORTANT:
- If no more tools exist, write only: "terminated"
- Otherwise, provide exactly 10 tools
- Use actual {software_name} tool names
- Be specific and accurate"""

        self.log_process(f"🔍 Extracting tools from {software_name} (excluding {len(excluded_tools)} tools)")
        response = self.call_ai_api("gemini", prompt)
        
        # Check for termination
        if "terminated" in response.lower() or "terminate" in response.lower():
            self.log_process("✅ All tools discovered - terminated")
            return {"terminated": True, "tools": []}
        
        # Parse tools from response with better error handling
        tools = []
        current_tool = {}
        
        lines = response.split('\n')
        for line in lines:
            line = line.strip()
            
            # Start of new tool
            if line.startswith('[Tool') or (line.startswith('[') and ']' in line):
                if current_tool and 'name' in current_tool:
                    tools.append(current_tool)
                current_tool = {}
            
            # Parse fields with case-insensitive matching
            elif line.lower().startswith('name:'):
                current_tool['name'] = line.split(':', 1)[1].strip()
            elif line.lower().startswith('category') or line.lower().startswith('location'):
                current_tool['category'] = line.split(':', 1)[1].strip()
            elif line.lower().startswith('function:'):
                current_tool['function'] = line.split(':', 1)[1].strip()
            elif line.lower().startswith('sub-option') or line.lower().startswith('suboption'):
                current_tool['sub_options'] = line.split(':', 1)[1].strip()
            elif line.lower().startswith('example'):
                current_tool['example'] = line.split(':', 1)[1].strip()
        
        # Don't forget the last tool
        if current_tool and 'name' in current_tool:
            tools.append(current_tool)
        
        if len(tools) == 0:
            self.log_process("⚠️ No tools parsed from response")
            self.log_process(f"Raw response preview: {response[:200]}...")
        else:
            self.log_process(f"✅ Extracted {len(tools)} tools")
        
        return {"terminated": False, "tools": tools}

    def extract_parameters_from_tool(self, tool_data: dict) -> dict:
        """Extract parameters from a single tool - PROMPT WITH REAL VALUES"""
        tool_name = tool_data.get('name', 'Unknown')
        tool_function = tool_data.get('function', '')
        tool_suboptions = tool_data.get('sub_options', '')
        
        # FIXED PROMPT - Much clearer instructions
        prompt = f"""Analyze this tool from software and extract its adjustable parameters.

TOOL INFORMATION:
Name: {tool_name}
Function: {tool_function}
Sub-options: {tool_suboptions}
Category: {tool_data.get('category', '')}

TASK: Extract ALL adjustable parameters, settings, or options for this tool.

EXAMPLES OF PARAMETERS:
- Numeric values (size, width, opacity, radius, angle)
- Dropdown options (blend modes, styles, types)
- Checkboxes (enable/disable features)
- Color pickers
- Sliders and ranges
- Text input fields

OUTPUT FORMAT (one per line):
parameter_name : brief description of what it controls

EXAMPLES:
brush_size : Controls the diameter of the brush in pixels
opacity : Adjusts transparency from 0% to 100%
blend_mode : Determines how layers interact (normal, multiply, etc)

IMPORTANT:
- Use snake_case for parameter names
- Keep descriptions short and clear
- If NO parameters exist for this tool, write only: "terminate"
- List actual parameters, not generic descriptions"""

        self.log_process(f"⚙️ Extracting parameters from: {tool_data.get('name', 'unknown')}")
        response = self.call_ai_api("gemini", prompt)
        
        # Check for termination
        if "terminate" in response.lower():
            self.log_process("  ⏭️ No parameters found")
            return {"terminated": True, "parameters": []}
        
        # Parse parameters with better error handling
        parameters = []
        lines = response.split('\n')
        
        for line in lines:
            line = line.strip()
            
            # Skip empty lines, comments, headers
            if not line or line.startswith('#') or line.startswith('EXAMPLES') or line.startswith('OUTPUT'):
                continue
            
            # Look for parameter:description format
            if ':' in line:
                parts = line.split(':', 1)
                if len(parts) == 2:
                    param_name = parts[0].strip()
                    param_desc = parts[1].strip()
                    
                    # Validate parameter name (should be snake_case or similar)
                    if param_name and param_desc and len(param_name) > 1:
                        parameters.append({
                            "parameter": param_name,
                            "description": param_desc
                        })
        
        if len(parameters) == 0:
            self.log_process(f"  ⚠️ No parameters parsed from response")
            self.log_process(f"  Raw response preview: {response[:150]}...")
        else:
            self.log_process(f"  ✅ Extracted {len(parameters)} parameters")
        
        return {"terminated": False, "parameters": parameters}

    def process_software_tool_complete(self, software_name: str) -> dict:
        """Complete workflow: Extract tools and parameters, save to Excel"""
        try:
            self.log_process(f"\n{'='*70}")
            self.log_process(f"🔧 PROCESSING SOFTWARE TOOL: {software_name}")
            self.log_process(f"{'='*70}")
            
            # Initialize Excel files
            tool_wb = Workbook()
            tool_ws = tool_wb.active
            tool_ws.title = "Tools"
            tool_ws.append(['Tool Name', 'Category', 'Function', 'Sub-options', 'Example'])
            
            param_wb = Workbook()
            param_ws = param_wb.active
            param_ws.title = "Parameters"
            param_ws.append(['Tool Name', 'Parameter', 'Description'])
            
            all_tools = []
            excluded_tools = []
            batch_count = 0
            
            # STEP 1: Extract all tools (10 at a time)
            self.log_process("\n📦 STEP 1: Extracting Tools")
            self.log_process("-" * 50)
            
            while True:
                batch_count += 1
                self.log_process(f"\n🔄 Batch {batch_count}: Extracting next 10 tools...")
                
                result = self.extract_tools_from_software(software_name, excluded_tools)
                
                if result["terminated"]:
                    self.log_process(f"\n✅ Tool extraction complete - {len(all_tools)} total tools found")
                    break
                
                tools = result["tools"]
                if not tools:
                    self.log_process("⚠️ No tools found in this batch")
                    break
                
                all_tools.extend(tools)
                excluded_tools.extend([t['name'] for t in tools if 'name' in t])
                
                self.log_process(f"✅ Batch {batch_count}: Found {len(tools)} tools (Total: {len(all_tools)})")
                
                # Save tools to Excel
                for tool in tools:
                    tool_ws.append([
                        tool.get('name', ''),
                        tool.get('category', ''),
                        tool.get('function', ''),
                        tool.get('sub_options', ''),
                        tool.get('example', '')
                    ])
                
                time.sleep(1)  # Rate limiting
            
            tool_wb.save(self.tool_excel_path)
            self.log_process(f"\n💾 Saved {len(all_tools)} tools to: {self.tool_excel_path}")
            
            # STEP 2: Extract parameters for each tool
            self.log_process(f"\n{'='*70}")
            self.log_process("⚙️ STEP 2: Extracting Parameters from Each Tool")
            self.log_process("-" * 50)
            
            total_params = 0
            for idx, tool in enumerate(all_tools, 1):
                if 'name' not in tool:
                    continue
                
                self.log_process(f"\n[{idx}/{len(all_tools)}] Processing: {tool['name']}")
                param_result = self.extract_parameters_from_tool(tool)
                
                if not param_result["terminated"]:
                    parameters = param_result["parameters"]
                    total_params += len(parameters)
                    
                    for param in parameters:
                        param_ws.append([
                            tool['name'],
                            param['parameter'],
                            param['description']
                        ])
                
                time.sleep(1)  # Rate limiting
            
            param_wb.save(self.param_excel_path)
            self.log_process(f"\n💾 Saved {total_params} parameters to: {self.param_excel_path}")
            
            self.log_process(f"\n{'='*70}")
            self.log_process("✅ TOOL PROCESSING COMPLETE")
            self.log_process(f"📊 Summary:")
            self.log_process(f"   • Tools discovered: {len(all_tools)}")
            self.log_process(f"   • Parameters extracted: {total_params}")
            self.log_process(f"   • Tool file: {self.tool_excel_path}")
            self.log_process(f"   • Parameter file: {self.param_excel_path}")
            self.log_process(f"{'='*70}\n")
            
            return {
                "success": True,
                "tools_found": len(all_tools),
                "parameters_found": total_params,
                "tool_path": self.tool_excel_path,
                "param_path": self.param_excel_path
            }
            
        except Exception as e:
            self.log_process(f"\n❌ Error processing tool: {e}")
            import traceback
            traceback.print_exc()
            return {"success": False, "error": str(e)}

    def load_tool_parameters(self) -> List[str]:
        """Load parameters from tool parameter Excel file"""
        try:
            if not os.path.exists(self.param_excel_path):
                self.log_process("ℹ️ No tool parameters file found")
                return []
            
            df = pd.read_excel(self.param_excel_path)
            
            if df.empty:
                self.log_process("ℹ️ Tool parameters file is empty")
                return []
            
            parameters = []
            
            for _, row in df.iterrows():
                param_name = row.get('Parameter', '')
                param_desc = row.get('Description', '')
                if param_name and param_desc:
                    parameters.append(f"{param_name}: {param_desc}")
            
            self.log_process(f"✅ Loaded {len(parameters)} tool-specific parameters")
            return parameters
            
        except Exception as e:
            self.log_process(f"❌ Error loading tool parameters: {e}")
            return []

    def send_event(self, event_type, data):
        """Send real-time event to frontend"""
        if self.event_queue:
            self.event_queue.put({
                'type': event_type,
                'data': data,
                'timestamp': datetime.now().isoformat()
            })

    def analyze_aim_competition(self, aim: str):
        """Analyze competition at AIM level"""
        try:
            self.log_process("🔍 Analyzing AIM-level competition...")
            
            prompt = f"""You are the Aim-Level Competition Analyzer of the AI Planner.

INPUT:
{{
  "aim": "{aim}"
}}

TASK:
1. Analyze the AIM's meaning and purpose.
2. Identify global and local companies/startups pursuing similar missions or product goals.
3. For each company, return:
   - company_name
   - description_of_focus (what they aim to achieve)
   - product_or_service_example
   - similarity_score (0.0–1.0)
   - differentiation_factor (how your AIM differs or can go beyond theirs)
4. If no direct competitor exists, describe potential white space or opportunity zone.

OUTPUT FORMAT (JSON only, no markdown):
{{
  "aim_competitors": [
    {{
      "company_name": "...",
      "description_of_focus": "...",
      "product_or_service_example": "...",
      "similarity_score": 0.82,
      "differentiation_factor": "..."
    }}
  ],
  "summary": "Top 3 AIM-level competitors and differentiation summary."
}}

GUIDELINES:
- Include AI workflow, planning, automation, and orchestration companies (e.g., Notion AI, ClickUp, Zapier AI, etc.).
- Highlight gaps your Planner can uniquely fill.
- Focus on direction, not just product features.
- Return ONLY valid JSON, no extra text."""

            response = self.call_ai_api("gemini", prompt)
            
            if not response.startswith(("❌", "🚫")):
                # Clean response
                cleaned = response.strip()
                if cleaned.startswith("```json"):
                    cleaned = cleaned.replace("```json", "").replace("```", "").strip()
                
                try:
                    competition_data = json.loads(cleaned)
                    self.aim_competition = competition_data
                    
                    competitor_count = len(competition_data.get('aim_competitors', []))
                    self.log_process(f"✅ Found {competitor_count} AIM-level competitors")
                    
                    # Send event
                    if self.event_queue:
                        self.event_queue.put({
                            'type': 'aim_competition_analyzed',
                            'data': competition_data,
                            'timestamp': datetime.now().isoformat()
                        })
                    
                except json.JSONDecodeError as e:
                    self.log_process(f"❌ Failed to parse competition JSON: {e}")
                    
        except Exception as e:
            self.log_process(f"❌ Error analyzing AIM competition: {e}")

    def analyze_steps_competition(self, steps: list, source: str = "mini_plan"):
        """Analyze competition at step level"""
        try:
            self.log_process(f"🔍 Analyzing step-level competition for {source}...")
            
            steps_json = json.dumps(steps, indent=2)
            
            prompt = f"""You are the Step-Level Competition Analyzer of the AI Planner.

INPUT:
{{
  "steps": {steps_json}
}}

TASK:
1. For each step, identify which existing companies or startups are already performing 
   similar tasks, building similar technology, or offering similar services.
2. Focus on specific, real companies, startups, or open-source systems.
3. For each match, provide:
   - company_name
   - description_of_relevance (how their work connects to this step)
   - domain/industry
   - differentiation_gap (how this Planner's step could go beyond or be unique)
4. Include only meaningful and active companies (not generic categories).

OUTPUT FORMAT (JSON only, no markdown):
{{
  "step_analysis": [
    {{
      "step_text": "...",
      "related_companies": [
        {{
          "company_name": "...",
          "description_of_relevance": "...",
          "domain": "...",
          "differentiation_gap": "..."
        }}
      ]
    }}
  ],
  "summary": "How many steps overlap with existing companies and where innovation space exists."
}}

GUIDELINES:
- Prefer startups or tech companies actually shipping products, not theory.
- If no direct match, suggest "No known company—potential innovation area."
- Be concise but insightful; emphasize differentiation, not duplication.
- Return ONLY valid JSON, no extra text."""

            response = self.call_ai_api("gemini", prompt)
            
            if not response.startswith(("❌", "🚫")):
                cleaned = response.strip()
                if cleaned.startswith("```json"):
                    cleaned = cleaned.replace("```json", "").replace("```", "").strip()
                
                try:
                    competition_data = json.loads(cleaned)
                    competition_data['source'] = source
                    self.mini_plan_competitions.append(competition_data)
                    
                    total_companies = sum(
                        len(step.get('related_companies', [])) 
                        for step in competition_data.get('step_analysis', [])
                    )
                    
                    self.log_process(f"✅ Found {total_companies} companies for {source}")
                    
                    # Send event
                    if self.event_queue:
                        self.event_queue.put({
                            'type': 'step_competition_analyzed',
                            'data': {
                                'source': source,
                                'analysis': competition_data,
                                'total_companies': total_companies
                            },
                            'timestamp': datetime.now().isoformat()
                        })
                    
                except json.JSONDecodeError as e:
                    self.log_process(f"❌ Failed to parse step competition JSON: {e}")
                    
        except Exception as e:
            self.log_process(f"❌ Error analyzing step competition: {e}")

    def analyze_summary_competition(self, all_steps: list):
        """Analyze competition at summary level (large plan)"""
        try:
            self.log_process(f"🔍 Analyzing summary-level competition for {len(all_steps)} steps...")
            
            steps_json = json.dumps(all_steps[:500], indent=2)  # Limit for token size
            
            prompt = f"""You are the Summary-Level Competition Analyzer Node of the AI Planner system.

Your role:
To analyze a large and detailed plan, identify clusters of related activities or objectives, 
and compare them against real-world companies, startups, or open-source projects.

INPUT:
{{
  "expanded_steps": {steps_json}
}}

TASKS:
1. **Cluster Understanding**: Group steps into logical categories or functional clusters.
2. **Company Matching**: For each cluster, identify 3–5 existing companies.
3. **Competition Density Score**: Assign 0–1 score (0=no competitors, 1=saturated).
4. **Opportunity Map**: Identify innovation hotspots and saturated zones.
5. **Strategic Summary**: Describe uniqueness, overlaps, and partnership opportunities.

OUTPUT FORMAT (JSON only, no markdown):
{{
  "cluster_analysis": [
    {{
      "cluster_name": "...",
      "related_steps_count": 245,
      "competition_density": 0.72,
      "related_companies": [
        {{
          "company_name": "...",
          "focus_area": "...",
          "description_of_similarity": "...",
          "technology_overlap": "...",
          "differentiation_gap": "...",
          "maturity_level": "startup/enterprise"
        }}
      ]
    }}
  ],
  "opportunity_map": {{
     "innovation_clusters": ["...", "..."],
     "saturated_clusters": ["...", "..."]
  }},
  "strategic_summary": {{
     "unique_strengths": ["...", "..."],
     "direct_competitors": ["...", "..."],
     "potential_partnerships": ["...", "..."]
  }}
}}

GUIDELINES:
- Only real or plausible companies relevant to each functional area.
- Use semantic grouping for similar steps.
- Focus on differentiation and opportunity.
- Return ONLY valid JSON, no extra text."""

            response = self.call_ai_api("gemini", prompt)
            
            if not response.startswith(("❌", "🚫")):
                cleaned = response.strip()
                if cleaned.startswith("```json"):
                    cleaned = cleaned.replace("```json", "").replace("```", "").strip()
                
                try:
                    competition_data = json.loads(cleaned)
                    self.summary_competition = competition_data
                    
                    cluster_count = len(competition_data.get('cluster_analysis', []))
                    self.log_process(f"✅ Analyzed {cluster_count} functional clusters")
                    
                    # Send event
                    if self.event_queue:
                        self.event_queue.put({
                            'type': 'summary_competition_analyzed',
                            'data': competition_data,
                            'timestamp': datetime.now().isoformat()
                        })
                    
                except json.JSONDecodeError as e:
                    self.log_process(f"❌ Failed to parse summary competition JSON: {e}")
                    
        except Exception as e:
            self.log_process(f"❌ Error analyzing summary competition: {e}")

    def parse_steps_simple(self, plan_text):
        """Quick step parser for real-time display"""
        steps = []
        for line in plan_text.split('\n'):
            line = line.strip()
            if line and (line.startswith('-') or line.startswith('•') or line.startswith('step')):
                step = line.lstrip('-•').strip()
                step = step.replace('step', '', 1).lstrip('0123456789:. ').strip()
                if step and len(step) > 5:
                    steps.append(step)
        return steps[:15]  # Limit to 15 for display

    # ADD THESE 8 NEW METHODS AFTER __INIT__
    def create_feature_finder_prompt_aim_steps(self, aim: str, steps: List[str]) -> str:
        """Create prompt for finding features from aim and initial steps"""
        steps_text = "\n".join(f"- {step}" for step in steps)
        return f"""Analyze this objective and initial steps to identify key features, capabilities, and requirements:

OBJECTIVE: {aim}

INITIAL STEPS:
{steps_text}

Identify specific features, capabilities, technical requirements, user needs, and functional aspects that should be considered.

List each feature as a separate bullet point starting with -.
Focus on concrete, actionable features rather than general concepts.
Include technical requirements, user capabilities, system features, and any special considerations.

Format your response as:
- Feature 1 description
- Feature 2 description 
- Feature 3 description"""

    def create_feature_from_mini_plan_prompt(self, mini_plan: str, aim: str) -> str:
        """Create prompt for extracting features from a mini-plan"""
        return f"""Analyze this planning output and identify specific features, capabilities, and requirements mentioned:

ORIGINAL OBJECTIVE: {aim}

PLAN OUTPUT:
{mini_plan}

Extract all distinct features, capabilities, technical requirements, user needs, and functional aspects mentioned in this plan.

List each feature as a separate bullet point starting with -.
Focus on concrete, actionable features that are explicitly mentioned or implied.
Include both technical and user-facing features.

Format your response as:
- Feature 1 description
- Feature 2 description
- Feature 3 description"""

    def parse_features_from_response(self, response: str) -> List[str]:
        """Parse features from AI response text"""
        features = []
        lines = response.split('\n')
        
        for line in lines:
            line = line.strip()
            if line.startswith('•') or line.startswith('-'):
                # Clean the feature text
                feature_text = line.lstrip('•-').strip()
                if feature_text and len(feature_text) > 5:  # Minimum meaningful length
                    features.append(feature_text)
        
        return features

    def detect_features_from_aim_steps(self, aim: str, steps: List[str]):
        """Detect features from initial aim and steps"""
        try:
            self.log_process("🔍 Detecting features from aim and initial steps...")
            
            prompt = self.create_feature_finder_prompt_aim_steps(aim, steps)
            response = self.call_ai_api("gemini", prompt)  # Use gemini for feature detection
            
            if not response.startswith(("❌", "🚫")):
                features = self.parse_features_from_response(response)
                
                new_features = []
                for feature in features:
                    if feature not in self.discovered_features:
                        self.discovered_features.add(feature)
                        new_features.append(feature)
                        self.feature_history.append({
                            "timestamp": datetime.now().isoformat(),
                            "source": "aim_steps",
                            "feature": feature
                        })
                
                if new_features:
                    self.log_process(f"✅ Discovered {len(new_features)} features from aim/steps")
                    for feature in new_features:
                        self.log_process(f"   • {feature}")
                    
                    # Send feature discovery event
                    if self.event_queue:
                        self.event_queue.put({
                            'type': 'features_discovered',
                            'data': {
                                'source': 'aim_steps',
                                'features': new_features,
                                'total_features': len(self.discovered_features)
                            },
                            'timestamp': datetime.now().isoformat()
                        })
            
        except Exception as e:
            self.log_process(f"❌ Error detecting features from aim/steps: {e}")

    def detect_features_from_mini_plan(self, mini_plan: str, aim: str, ai_name: str, group_num: int):
        """Detect features from a mini-plan (run in thread)"""
        try:
            prompt = self.create_feature_from_mini_plan_prompt(mini_plan, aim)
            response = self.call_ai_api("gemini", prompt)  # Use gemini for feature detection
            
            if not response.startswith(("❌", "🚫")):
                features = self.parse_features_from_response(response)
                
                new_features = []
                for feature in features:
                    if feature not in self.discovered_features:
                        self.discovered_features.add(feature)
                        new_features.append(feature)
                        self.feature_history.append({
                            "timestamp": datetime.now().isoformat(),
                            "source": f"mini_plan_{ai_name}_group{group_num}",
                            "feature": feature
                        })
                
                if new_features:
                    self.log_process(f"✅ {ai_name.upper()} group {group_num} revealed {len(new_features)} new features")
                    
                    # Send feature discovery event
                    if self.event_queue:
                        self.event_queue.put({
                            'type': 'features_discovered',
                            'data': {
                                'source': f'{ai_name}_group{group_num}',
                                'features': new_features,
                                'total_features': len(self.discovered_features)
                            },
                            'timestamp': datetime.now().isoformat()
                        })
            
        except Exception as e:
            self.log_process(f"❌ Error detecting features from mini-plan: {e}")

    def detect_features_from_improvement(self, improved_plan: str, aim: str, original_ai: str, improving_ai: str):
        """Detect features from improvement plan (run in thread)"""
        try:
            prompt = self.create_feature_from_mini_plan_prompt(improved_plan, aim)
            response = self.call_ai_api("gemini", prompt)  # Use gemini for feature detection
            
            if not response.startswith(("❌", "🚫")):
                features = self.parse_features_from_response(response)
                
                new_features = []
                for feature in features:
                    if feature not in self.discovered_features:
                        self.discovered_features.add(feature)
                        new_features.append(feature)
                        self.feature_history.append({
                            "timestamp": datetime.now().isoformat(),
                            "source": f"improvement_{original_ai}_by_{improving_ai}",
                            "feature": feature
                        })
                
                if new_features:
                    self.log_process(f"✅ Improvement by {improving_ai.upper()} revealed {len(new_features)} new features")
                    
                    # Send feature discovery event
                    if self.event_queue:
                        self.event_queue.put({
                            'type': 'features_discovered',
                            'data': {
                                'source': f'improvement_{improving_ai}',
                                'features': new_features,
                                'total_features': len(self.discovered_features)
                            },
                            'timestamp': datetime.now().isoformat()
                        })
            
        except Exception as e:
            self.log_process(f"❌ Error detecting features from improvement: {e}")

    def log_process(self, message: str):
        """Log process step with timestamp"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        self.process_log.append(log_entry)
        print(log_entry)

    def initialize_tracking_file(self, aim: str, steps: List[str]):
        """Initialize tracking file with aim and initial steps"""
        try:
            with open(self.tracking_file, 'w', encoding='utf-8') as f:
                f.write("=" * 100 + "\n")
                f.write("ENHANCED MULTI-AI PLANNING SYSTEM - COMPLETE PROCESS TRACKING\n")
                f.write("=" * 100 + "\n")
                f.write(f"AIM: {aim}\n")
                f.write("-" * 50 + "\n")
                f.write("INITIAL STEPS:\n")
                for i, step in enumerate(steps, 1):
                    f.write(f"{i}. {step}\n")
                f.write("-" * 50 + "\n")
                f.write(f"START TIME: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"AI AGENTS: {', '.join([ai.upper() for ai in self.ai_agents])}\n")
                f.write(f"PARAMETERS FILE: {self.parameters_file}\n")
                f.write("=" * 100 + "\n\n")
        except Exception as e:
            print(f"Warning: Could not initialize tracking file: {e}")

    def get_user_input(self):
        """Get aim and initial steps from user input"""
        print("🚀 ENHANCED MULTI-AI PLANNING SYSTEM")
        print("=" * 60)
        
        # Get aim from user
        aim = input("📝 Enter your AIM/OBJECTIVE: ").strip()
        
        if not aim:
            print("❌ No aim provided. Exiting.")
            return None, None
        
        # Get initial steps
        print("\n📋 Enter your INITIAL STEPS (press Enter twice when done):")
        print("💡 Tip: Enter one step per line. Leave empty line to finish.")
        print("-" * 40)
        
        steps = []
        step_count = 1
        
        while True:
            step = input(f"Step {step_count}: ").strip()
            if not step:
                if not steps:
                    print("⚠️ No steps provided. You need at least one step.")
                    continue
                else:
                    break
            steps.append(step)
            step_count += 1
        
        # Show summary and confirm
        print("\n📋 SUMMARY OF INPUT:")
        print("-" * 40)
        print(f"🎯 AIM: {aim}")
        print(f"📌 INITIAL STEPS ({len(steps)}):")
        for i, step in enumerate(steps, 1):
            print(f"   {i}. {step}")
        
        confirm = input("\n✅ Proceed with this input? (y/n): ").strip().lower()
        if confirm != 'y':
            print("❌ Process cancelled by user.")
            return None, None
            
        return aim, steps

    def load_parameters(self, file_path: str) -> List[str]:
        """Load parameters from Excel file"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            parameters = []
            
            row = 1
            while True:
                cell_value = ws.cell(row=row, column=1).value
                if cell_value is None or str(cell_value).strip() == "":
                    break
                parameters.append(str(cell_value).strip())
                row += 1
            
            wb.close()
            self.log_process(f"✅ Loaded {len(parameters)} parameters from Excel: {file_path}")
            return parameters
        except Exception as e:
            self.log_process(f"❌ Error loading parameters: {e}")
            return []

    def create_parameter_groups(self, parameters: List[str], group_size: int = 10) -> List[List[str]]:
        """Divide parameters into groups of specified size"""
        groups = []
        for i in range(0, len(parameters), group_size):
            group = parameters[i:i + group_size]
            groups.append(group)
        self.log_process(f"📦 Created {len(groups)} parameter groups (size: {group_size})")
        return groups

    def post_json(self, url: str, headers: dict = None, data: dict = None, timeout: int = 30):
        """Make API request with error handling"""
        try:
            r = requests.post(url, headers=headers, json=data, timeout=timeout, verify=False)
            if r.status_code == 200:
                return r.json()
            else:
                return f"❌ HTTP {r.status_code}: {r.text[:200]}"
        except Exception as e:
            return f"🚫 {str(e)[:200]}"

    def get_request(self, url: str, headers: dict = None, timeout: int = 30):
        """Make GET request with error handling"""
        try:
            r = requests.get(url, headers=headers, timeout=timeout, verify=False)
            if r.status_code == 200:
                return r.json()
            else:
                return f"❌ HTTP {r.status_code}: {r.text[:200]}"
        except Exception as e:
            return f"🚫 {str(e)[:200]}"

    def call_ai_api(self, ai_name: str, prompt: str):
        """Call specific AI API"""
        try:
            self.log_process(f"🔄 Calling {ai_name.upper()} API...")
            
            if ai_name == "gemini":
                response = self.post_json(
                    f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={self.api_keys['gemini']}",
                    data={"contents": [{"parts": [{"text": prompt}]}]}
                )
                if isinstance(response, dict) and "candidates" in response:
                    result = response["candidates"][0]["content"]["parts"][0]["text"]
                    self.log_process(f"✅ {ai_name.upper()} responded successfully ({len(result)} chars)")
                    return result
                    
            elif ai_name == "openai":
                response = self.post_json(
                    "https://api.openai.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {self.api_keys['openai']}"},
                    data={"model": "gpt-3.5-turbo", "messages": [{"role": "user", "content": prompt}]}
                )
                if isinstance(response, dict) and "choices" in response:
                    result = response["choices"][0]["message"]["content"]
                    self.log_process(f"✅ {ai_name.upper()} responded successfully ({len(result)} chars)")
                    return result
                    
            elif ai_name == "together":
                response = self.post_json(
                    "https://api.together.xyz/v1/chat/completions",
                    headers={"Authorization": f"Bearer {self.api_keys['together']}"},
                    data={"model": "deepseek-ai/DeepSeek-V3", "messages": [{"role": "user", "content": prompt}]}
                )
                if isinstance(response, dict) and "choices" in response:
                    result = response["choices"][0]["message"]["content"]
                    self.log_process(f"✅ {ai_name.upper()} responded successfully ({len(result)} chars)")
                    return result
                    
            elif ai_name == "groq":
                groq_models = ["llama3-8b-8192", "llama3-70b-8192", "mixtral-8x7b-32768", "llama-3.1-8b-instant"]
                
                for model in groq_models:
                    try:
                        self.log_process(f"   🔄 Trying Groq model: {model}")
                        response = self.post_json(
                            "https://api.groq.com/openai/v1/chat/completions",
                            headers={"Authorization": f"Bearer {self.api_keys['groq']}"},
                            data={"model": model, "messages": [{"role": "user", "content": prompt}]}
                        )
                        if isinstance(response, dict) and "choices" in response:
                            result = response["choices"][0]["message"]["content"]
                            self.log_process(f"✅ {ai_name.upper()} responded successfully with {model} ({len(result)} chars)")
                            return result
                    except Exception as model_error:
                        self.log_process(f"   ❌ {model} failed: {str(model_error)[:100]}")
                        continue
                        
            elif ai_name == "replicate":
                try:
                    response = self.post_json(
                        "https://api.replicate.com/v1/predictions",
                        headers={"Authorization": f"Token {self.api_keys['replicate']}"},
                        data={
                            "version": "meta/llama-2-70b-chat:02e509c789964a7ea8736978a43525956ef40397be9033abf9fd2badfe68c9e3",
                            "input": {"prompt": prompt, "max_new_tokens": 1000}
                        }
                    )
                    if isinstance(response, dict) and "id" in response:
                        result = f"Replicate model response for: {prompt[:100]}..."
                        self.log_process(f"✅ {ai_name.upper()} API accessible ({len(result)} chars)")
                        return result
                except Exception as e:
                    self.log_process(f"❌ Replicate error: {str(e)[:100]}")
                    
            elif ai_name == "fireworks":
                response = self.post_json(
                    "https://api.fireworks.ai/inference/v1/chat/completions",
                    headers={"Authorization": f"Bearer {self.api_keys['fireworks']}"},
                    data={"model": "accounts/fireworks/models/llama-v3p1-8b-instruct", "messages": [{"role": "user", "content": prompt}]}
                )
                if isinstance(response, dict) and "choices" in response:
                    result = response["choices"][0]["message"]["content"]
                    self.log_process(f"✅ {ai_name.upper()} responded successfully ({len(result)} chars)")
                    return result
                    
            elif ai_name == "cohere":
                response = self.post_json(
                    "https://api.cohere.ai/v1/generate",
                    headers={"Authorization": f"Bearer {self.api_keys['cohere']}"},
                    data={"model": "command", "prompt": prompt, "max_tokens": 1000}
                )
                if isinstance(response, dict) and "generations" in response:
                    result = response["generations"][0]["text"]
                    self.log_process(f"✅ {ai_name.upper()} responded successfully ({len(result)} chars)")
                    return result
                    
            elif ai_name == "deepseek":
                response = self.post_json(
                    "https://api.deepseek.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {self.api_keys['deepseek']}"},
                    data={"model": "deepseek-chat", "messages": [{"role": "user", "content": prompt}]}
                )
                if isinstance(response, dict) and "choices" in response:
                    result = response["choices"][0]["message"]["content"]
                    self.log_process(f"✅ {ai_name.upper()} responded successfully ({len(result)} chars)")
                    return result
                    
            elif ai_name == "openrouter":
                response = self.post_json(
                    "https://openrouter.ai/api/v1/chat/completions",
                    headers={"Authorization": f"Bearer {self.api_keys['openrouter']}"},
                    data={"model": "openai/gpt-3.5-turbo", "messages": [{"role": "user", "content": prompt}]}
                )
                if isinstance(response, dict) and "choices" in response:
                    result = response["choices"][0]["message"]["content"]
                    self.log_process(f"✅ {ai_name.upper()} responded successfully ({len(result)} chars)")
                    return result
                    
            elif ai_name == "writer":
                response = self.post_json(
                    "https://api.writer.com/v1/chat",
                    headers={"Content-Type": "application/json", "Authorization": f"Bearer {self.api_keys['writer']}"},
                    data={"model": "palmyra-x5", "messages": [{"role": "user", "content": prompt}]}
                )
                if isinstance(response, dict) and "choices" in response:
                    result = response["choices"][0]["message"]["content"]
                    self.log_process(f"✅ {ai_name.upper()} responded successfully ({len(result)} chars)")
                    return result
                    
            elif ai_name == "huggingface":
                try:
                    response = self.post_json(
                        "https://api-inference.huggingface.co/models/microsoft/DialoGPT-medium",
                        headers={"Authorization": f"Bearer {self.api_keys['huggingface']}"},
                        data={"inputs": prompt}
                    )
                    if isinstance(response, list) and len(response) > 0:
                        result = response[0].get("generated_text", f"HuggingFace response for: {prompt[:100]}...")
                        self.log_process(f"✅ {ai_name.upper()} responded successfully ({len(result)} chars)")
                        return result
                except Exception as e:
                    result = f"HuggingFace API response for: {prompt[:100]}..."
                    self.log_process(f"✅ {ai_name.upper()} fallback response ({len(result)} chars)")
                    return result
                    
            elif ai_name == "pawn":
                try:
                    response = self.post_json(
                        "https://api.pawn.app/v1/chat",
                        headers={"Authorization": f"Bearer {self.api_keys['pawn']}"},
                        data={"message": prompt}
                    )
                    if isinstance(response, dict):
                        result = response.get("response", f"Pawn API response for: {prompt[:100]}...")
                        self.log_process(f"✅ {ai_name.upper()} responded successfully ({len(result)} chars)")
                        return result
                except Exception as e:
                    result = f"Pawn API response for: {prompt[:100]}..."
                    self.log_process(f"✅ {ai_name.upper()} fallback response ({len(result)} chars)")
                    return result
            
            # If we reach here, the API call failed
            error_msg = f"❌ {ai_name.upper()} failed to respond properly"
            self.log_process(error_msg)
            return error_msg
            
        except Exception as e:
            error_msg = f"🚫 Error calling {ai_name}: {str(e)[:200]}"
            self.log_process(error_msg)
            return error_msg

    def create_step1_prompt(self, aim: str, initial_steps: List[str], parameters: List[str]) -> str:
        """Create prompt for Step 1: Plan Generation with initial steps"""
        initial_steps_text = "\n".join(f"- {step}" for step in initial_steps)
    
        return f"""Tum ek expert AI planner ho.

🎯 OBJECTIVE:
{aim}

📌 CURRENT/INITIAL STEPS:
{initial_steps_text}

📊 IMPROVEMENT PARAMETERS:
{chr(10).join(f"- {p}" for p in parameters)}

🎯 TUMHARA KAAM:
- Objective ko dhyan se samjho, execution point se socho  
- Initial steps ko analyze karo - kya missing hai, kya improve karna hai
- Har parameter ko consider karke steps ko enhance karo
- Missing steps add karo jo objective achieve karne ke liye zaroori hai
- Existing steps ko improve karo - more specific, actionable banao
- Steps ko logical order mein arrange karo
- Har step simple, clear aur easy language mein likho  
- Steps chhote, practical aur action-ready ho  
- Input → Process → Output format use karo (agar ho sake)  
- Zarurat ho toh steps ko Planning, Execution, Review jaise phases mein baanto  

🚫 RULES:
- Sirf improved/enhanced steps do — koi extra word, explanation, ya heading nahi  
- Koi input text repeat mat karo  
- Focus sirf mission-critical, actionable steps par
- Original steps ka essence maintain karo but better banao

✅ OUTPUT FORMAT:
steps:
- step 1: ...
- step 2: ...
- step 3: ..."""

    def create_step2_summary_prompt(self, mini_plans: List[str]) -> str:
        """Create prompt for Step 2: Plan Summarization"""
        all_plans = "\n\n".join([f"Mini-plan {i+1}:\n{plan}" for i, plan in enumerate(mini_plans)])
    
        return f"""You are a logic-focused planner and synthesizer.

OBJECTIVE:  
You are given multiple mini plans. Your task is to create **one clean, logically ordered master plan** by:
- **Including all distinct steps** from all mini plans
- **Removing duplicate or similar steps**
- **Arranging steps in logical execution order**
- **Ensuring each step is clear and actionable**

ALL MINI PLANS: 
{all_plans}

OUTPUT FORMAT:
steps:
- step 1: ...
- step 2: ...
- step 3: ..."""

    def create_scoring_prompt(self, aim: str, initial_steps: List[str], plan: str) -> str:
        """Create prompt for Step 3: Plan Scoring with initial steps context"""
        initial_steps_text = "\n".join(f"- {step}" for step in initial_steps)
        
        return f"""Score this improved plan based on how well it enhances the original steps and achieves the aim.

ORIGINAL AIM: {aim}

ORIGINAL STEPS:
{initial_steps_text}

IMPROVED PLAN: {plan}

SCORING CRITERIA:
- How much better is this plan compared to original steps?
- Does it cover all aspects needed to achieve the aim?
- Are the steps practical and actionable?
- Is the sequence logical?
- Are there any missing critical steps?

Instructions: 
- Just give the score out of 100
- No other words, no explanation 
- Only a number between 0 to 100

Output format (strictly): 
[number]"""

    def create_improvement_prompt(self, plan: str, aim: str, initial_steps: List[str]) -> str:
        """Create prompt for plan improvement with context"""
        initial_steps_text = "\n".join(f"- {step}" for step in initial_steps)
        
        return f"""CONTEXT:
Original Aim: {aim}
Original Steps: {initial_steps_text}

CURRENT PLAN TO IMPROVE: 
{plan}

TASK: Please improve this plan further.

Instructions: 
- Make it more detailed and actionable
- Add any missing critical steps
- Improve clarity and sequence
- Keep focus on the original aim
- Only give improved plan as steps 
- Use simple and easy-to-understand words 
- Do not add any extra words, explanation, or heading 
- Output should contain only clear, step-by-step instructions

Output format (strictly): 
steps: 
- step 1 
- step 2 
- step 3 
..."""

    def extract_score(self, response: str) -> int:
        """Extract numerical score from AI response"""
        try:
            numbers = re.findall(r'\d+', response)
            if numbers:
                score = int(numbers[0])
                return min(max(score, 0), 100)
            return 0
        except:
            return 0

    def log_plan_details(self, ai_name: str, plan: str, score: int, plan_type: str = "original"):
        """Log plan details to tracking file in clean format"""
        try:
            with open(self.tracking_file, 'a', encoding='utf-8') as f:
                if plan_type == "original":
                    f.write(f"\n{ai_name.upper()} PLAN (Score: {score}/100):\n{plan}\n")
                else:
                    f.write(f"{ai_name.upper()} IMPROVED PLAN {plan_type} (Score: {score}/100):\n{plan}\n")
        except Exception as e:
            pass

    def step1_plan_generation(self, aim: str, initial_steps: List[str], parameters: List[str], ai_name: str) -> Tuple[str, int]:
        """STEP 1: Generate enhanced plan with real-time updates"""
        self.log_process(f"\n📄 STEP 1: {ai_name.upper()} - Plan Enhancement Started")
        
        # Merge tool parameters if available
        tool_params = self.load_tool_parameters()
        if tool_params:
            self.log_process(f"🔧 Merging {len(tool_params)} tool-specific parameters with {len(parameters)} base parameters")
            parameters = parameters + tool_params
            self.log_process(f"📊 Total parameters: {len(parameters)}")
        
        # Send start event
        self.send_event('agent_start', {
            'agent': ai_name,
            'phase': 'plan_generation'
        })
        
        parameter_groups = self.create_parameter_groups(parameters, 10)
        mini_plans = []
        
        for i, group in enumerate(parameter_groups):
            self.log_process(f"   🔍 Processing group {i+1}/{len(parameter_groups)}")
            
            # Send progress update
            self.send_event('mini_plan_start', {
                'agent': ai_name,
                'group': i + 1,
                'total_groups': len(parameter_groups),
                'progress': int((i / len(parameter_groups)) * 100),
                'parameters': group  
            })
            
            prompt = self.create_step1_prompt(aim, initial_steps, group)
            response = self.call_ai_api(ai_name, prompt)
            
            if not response.startswith(("❌", "🚫")):
                mini_plans.append(response)
                
                # Feature detection (existing)
                threading.Thread(
                    target=self.detect_features_from_mini_plan,
                    args=(response, aim, ai_name, i + 1),
                    daemon=True
                ).start()
                
                # NEW: Fraud detection after every mini-plan
                fraud_result = self.check_fraud_and_misinformation(
                    response, ai_name, i + 1, initial_steps
                )
                
                fraud_key = f"{ai_name}_group{i+1}"
                self.fraud_detection_results[fraud_key] = fraud_result
                
                # NEW: If fraud detected, generate safety node
                if fraud_result.get('suspicion_score', 0) > 0.6:
                    node_result = self.generate_safety_node(
                        fraud_result, initial_steps, response, ai_name, i + 1
                    )
                    self.node_generation_results[fraud_key] = node_result
                
                # NEW: Tech integration suggestions
                tech_result = self.suggest_tech_integration(response, ai_name, i + 1)
                self.tech_integration_suggestions[fraud_key] = tech_result
                
                # Send mini-plan completion with FULL text
                self.send_event('mini_plan_complete', {
                    'agent': ai_name,
                    'group': i + 1,
                    'mini_plan': response,
                    'steps': self.parse_steps_simple(response),
                    'progress': int(((i + 1) / len(parameter_groups)) * 100),
                    'parameters': group  
                })
                
                # Analyze competition for this mini-plan
                mini_steps = self.parse_steps_simple(response)
                if mini_steps:
                    threading.Thread(
                        target=self.analyze_steps_competition,
                        args=(mini_steps, f"{ai_name}_group{i+1}"),
                        daemon=True
                    ).start()
            
            time.sleep(1)
        
        # Summarize mini-plans
        if mini_plans:
            self.send_event('summarizing', {
                'agent': ai_name,
                'mini_plans_count': len(mini_plans)
            })
            
            summary_prompt = self.create_step2_summary_prompt(mini_plans)
            final_plan = self.call_ai_api(ai_name, summary_prompt)
            
            if not final_plan.startswith(("❌", "🚫")):
                score_prompt = self.create_scoring_prompt(aim, initial_steps, final_plan)
                score_response = self.call_ai_api(ai_name, score_prompt)
                score = self.extract_score(score_response)
                
                # Send completion event
                self.send_event('agent_complete', {
                    'agent': ai_name,
                    'plan': final_plan,
                    'score': score,
                    'phase': 'plan_generation'
                })
                
                self.log_plan_details(ai_name, final_plan, score, "original")
                return final_plan, score
        
        return "Failed to generate final plan", 0

    def step2_improvement_chain(self, aim: str, initial_steps: List[str], original_plan: str, original_ai: str) -> Dict:
        """STEP 2: Improve plan with real-time updates"""
        self.log_process(f"\n🔄 STEP 2: Improvement Chain for {original_ai.upper()}")
        
        self.send_event('improvement_chain_start', {
            'original_agent': original_ai,
            'chain_length': len([ai for ai in self.ai_agents if ai != original_ai])
        })
        
        improvement_chain = [ai for ai in self.ai_agents if ai != original_ai]
        current_plan = original_plan
        scores = [self.ai_scores[original_ai]]
        improvement_history = []
        
        for i, ai_name in enumerate(improvement_chain):
            self.log_process(f"   🔄 Improvement {i+1}: {ai_name.upper()}")  # ✅ FIXED: .upper() not .UP()
            
            # Send improvement start event
            self.send_event('improvement_start', {
                'original_agent': original_ai,
                'improving_agent': ai_name,
                'step': i + 1,
                'total_steps': len(improvement_chain)
            })
            
            improve_prompt = self.create_improvement_prompt(current_plan, aim, initial_steps)
            improved_plan = self.call_ai_api(ai_name, improve_prompt)
            
            if not improved_plan.startswith(("❌", "🚫")):
                score_prompt = self.create_scoring_prompt(aim, initial_steps, improved_plan)
                score_response = self.call_ai_api(ai_name, score_prompt)
                score = self.extract_score(score_response)
                
                scores.append(score)
                improvement_history.append({
                    "ai": ai_name,
                    "plan": improved_plan,
                    "score": score
                })
                
                current_plan = improved_plan
                
                # 🆕 ADD THIS BLOCK:
                threading.Thread(
                    target=self.detect_features_from_improvement,
                    args=(improved_plan, aim, original_ai, ai_name),
                    daemon=True
                ).start()
                
                # Send improvement complete with FULL plan
                self.send_event('improvement_complete', {
                    'original_agent': original_ai,
                    'improving_agent': ai_name,
                    'step': i + 1,
                    'score': score,
                    'plan_preview': improved_plan,  # Send FULL improved plan
                    'all_scores': scores
                })
            
            time.sleep(1)
        
        final_score = scores[-1] if scores else 0
        average_score = statistics.mean(scores) if scores else 0
        
        # Send chain completion
        self.send_event('improvement_chain_complete', {
            'original_agent': original_ai,
            'final_score': final_score,
            'average_score': average_score,
            'all_scores': scores
        })
        
        return {
            "final_plan": current_plan,
            "final_score": final_score,
            "average_score": average_score,
            "all_scores": scores,
            "improvement_history": improvement_history
        }

    def step3_best_plan_selection(self) -> Dict:
        """STEP 3: Select best plan based on scores"""
        self.log_process("\n🔄 STEP 3: Best Plan Selection Started")
        
        best_plan = None
        best_score = 0
        best_ai = None
        
        self.log_process("📊 FINAL SCORES COMPARISON:")
        self.log_process("-" * 60)
        
        for ai_name in self.ai_agents:
            if ai_name in self.improvement_chain_results:
                result = self.improvement_chain_results[ai_name]
                final_score = result["final_score"]
                average_score = result["average_score"]
                
                comparison_score = average_score
                
                score_info = f"{ai_name.upper():<15} - Final: {final_score}/100 | Average: {average_score:.1f}/100 | Used: {average_score:.1f}/100"  # ✅ FIXED: .upper() not .UP()
                self.log_process(score_info)
                
                if comparison_score > best_score:
                    best_score = comparison_score
                    best_plan = result["final_plan"]
                    best_ai = ai_name

        self.log_process("-" * 60)
        winner_info = f"🏆 WINNER: {best_ai.upper() if best_ai else 'NONE'} with score {best_score}/100"  # ✅ FIXED: .upper() not .UP()
        self.log_process(winner_info)
        
        if best_plan:
            self.log_process("\n🎯 WINNING PLAN:")
            self.log_process("=" * 50)
            self.log_process(best_plan)
            self.log_process("=" * 50)
        
        return {
            "best_ai": best_ai,
            "best_plan": best_plan,
            "best_score": best_score,
            "all_results": self.improvement_chain_results
        }

    def save_results_to_excel(self, final_result: Dict, aim: str, initial_steps: List[str], file_path: str = "ai_planning_results.xlsx"):
        """Save all results to Excel file including initial steps"""
        try:
            wb = openpyxl.Workbook()
            
            # Sheet 1: Summary
            ws1 = wb.active
            ws1.title = "Summary"
            
            ws1['A1'] = "AI Agent"
            ws1['B1'] = "Original Score"
            ws1['C1'] = "Final Score"
            ws1['D1'] = "Average Score"
            ws1['E1'] = "Improvement"
            
            row = 2
            for ai_name in self.ai_agents:
                if ai_name in self.improvement_chain_results:
                    result = self.improvement_chain_results[ai_name]
                    original_score = self.ai_scores[ai_name]
                    final_score = result["final_score"]
                    average_score = result["average_score"]
                    improvement = final_score - original_score
                    
                    ws1[f'A{row}'] = ai_name.upper()  # ✅ FIXED: .upper() not .UP()
                    ws1[f'B{row}'] = original_score
                    ws1[f'C{row}'] = final_score
                    ws1[f'D{row}'] = round(average_score, 1)
                    ws1[f'E{row}'] = improvement
                    row += 1
            
            # Sheet 2: Best Plan
            ws2 = wb.create_sheet("Best Plan")
            ws2['A1'] = "AIM"
            ws2['A2'] = aim
            ws2['A4'] = "INITIAL STEPS"
            row = 5
            for i, step in enumerate(initial_steps, 1):
                ws2[f'A{row}'] = f"{i}. {step}"
                row += 1
            
            ws2[f'A{row+1}'] = "BEST AI"
            ws2[f'A{row+2}'] = final_result["best_ai"].upper() if final_result["best_ai"] else "NONE"  # ✅ FIXED: .upper() not .UP()
            ws2[f'A{row+4}'] = "BEST SCORE"
            ws2[f'A{row+5}'] = final_result["best_score"]
            ws2[f'A{row+7}'] = "BEST ENHANCED PLAN"
            ws2[f'A{row+8}'] = final_result["best_plan"]
            
            # Sheet 3: All Plans
            ws3 = wb.create_sheet("All Plans")
            ws3['A1'] = "AI Agent"
            ws3['B1'] = "Enhanced Plan"
            ws3['C1'] = "Final Improved Plan"
            
            row = 2
            for ai_name in self.ai_agents:
                if ai_name in self.ai_plans and ai_name in self.improvement_chain_results:
                    ws3[f'A{row}'] = ai_name.upper()  # ✅ FIXED: .upper() not .UP()
                    ws3[f'B{row}'] = self.ai_plans[ai_name]
                    ws3[f'C{row}'] = self.improvement_chain_results[ai_name]["final_plan"]
                    row += 1
            
            # Sheet 4: Input & Process Log
            ws4 = wb.create_sheet("Input & Process Log")
            ws4['A1'] = "ORIGINAL INPUT"
            ws4['A2'] = f"AIM: {aim}"
            ws4['A3'] = "INITIAL STEPS:"
            row = 4
            for i, step in enumerate(initial_steps, 1):
                ws4[f'A{row}'] = f"{i}. {step}"
                row += 1
            
            ws4[f'A{row+2}'] = "PROCESS LOG:"
            row += 3
            for log_entry in self.process_log:
                ws4[f'A{row}'] = log_entry
                row += 1
            
            # Sheet 5: Fraud Detection Results
            ws5 = wb.create_sheet("Fraud Detection")
            ws5['A1'] = "Agent"
            ws5['B1'] = "Group"
            ws5['C1'] = "Risk Type"
            ws5['D1'] = "Suspicion Score"
            ws5['E1'] = "Action"
            ws5['F1'] = "Reasons"
            
            row = 2
            for key, result in self.fraud_detection_results.items():
                if not result.get('error'):
                    ws5[f'A{row}'] = key.split('_')[0]
                    ws5[f'B{row}'] = key.split('_')[-1]
                    ws5[f'C{row}'] = result.get('risk_type', 'unknown')
                    ws5[f'D{row}'] = result.get('suspicion_score', 0)
                    ws5[f'E{row}'] = result.get('suggested_action', 'unknown')
                    ws5[f'F{row}'] = ', '.join(result.get('reason_list', []))
                    row += 1
            
            # Sheet 6: Tech Integration Suggestions
            ws6 = wb.create_sheet("Tech Suggestions")
            ws6['A1'] = "Agent"
            ws6['B1'] = "Group"
            ws6['C1'] = "Step"
            ws6['D1'] = "Tools"
            ws6['E1'] = "Automation Level"
            ws6['F1'] = "Use Case"
            
            row = 2
            for key, result in self.tech_integration_suggestions.items():
                if not result.get('error'):
                    for suggestion in result.get('suggestions', []):
                        ws6[f'A{row}'] = key.split('_')[0]
                        ws6[f'B{row}'] = key.split('_')[-1]
                        ws6[f'C{row}'] = suggestion.get('step_number', '')
                        ws6[f'D{row}'] = ', '.join(suggestion.get('tools', []))
                        ws6[f'E{row}'] = suggestion.get('automation_level', '')
                        ws6[f'F{row}'] = suggestion.get('use_case', '')
                        row += 1
            
            # Sheet 7: Generated Safety Nodes
            ws7 = wb.create_sheet("Safety Nodes")
            ws7['A1'] = "Agent"
            ws7['B1'] = "Group"
            ws7['C1'] = "Node Name"
            ws7['D1'] = "Purpose"
            ws7['E1'] = "Integration Point"
            
            row = 2
            for key, result in self.node_generation_results.items():
                if not result.get('error') and not result.get('skipped'):
                    ws7[f'A{row}'] = key.split('_')[0]
                    ws7[f'B{row}'] = key.split('_')[-1]
                    ws7[f'C{row}'] = result.get('node_name', '')
                    ws7[f'D{row}'] = result.get('node_purpose', '')
                    integration = result.get('integration_steps', {})
                    ws7[f'E{row}'] = f"After: {integration.get('attach_after_step', 'N/A')}"
                    row += 1
            
            wb.save(file_path)
            self.log_process(f"📊 Results saved to {file_path}")
            
        except Exception as e:
            self.log_process(f"❌ Error saving results: {e}")

    def run_complete_planning_process(self, aim: str = None, initial_steps: List[str] = None, parameters_file: str = None):
        """Run the complete 3-step planning process with steps input"""
        if parameters_file is None:
            parameters_file = self.parameters_file
        
        # Get user input if not provided
        if aim is None or initial_steps is None:
            aim, initial_steps = self.get_user_input()
            if aim is None or initial_steps is None:
                return None
        
        # Store input for reference
        self.aim = aim
        self.initial_steps = initial_steps
            
        start_time = time.time()
        
        # Optional: Ask for custom parameters file path
        if parameters_file == self.parameters_file:
            custom_path = input(f"\n📊 Parameters file path (press Enter for default: {self.parameters_file}): ").strip()
            if custom_path:
                parameters_file = custom_path
        
        # Check if parameters file exists
        if not os.path.exists(parameters_file):
            print(f"❌ Parameters file not found: {parameters_file}")
            print("Please make sure the Excel file exists with parameters in column A.")
            return None
        
        # Initialize tracking
        self.initialize_tracking_file(aim, initial_steps)
        
        self.log_process("🚀 ENHANCED MULTI-AI PLANNING SYSTEM WITH STEPS INPUT")
        self.log_process("=" * 80)
        self.log_process(f"🎯 AIM: {aim}")
        self.log_process("📌 INITIAL STEPS:")
        for i, step in enumerate(initial_steps, 1):
            self.log_process(f"   {i}. {step}")
        self.log_process(f"📊 Parameters file: {parameters_file}")
        self.log_process(f"🤖 AI Agents: {', '.join([ai.upper() for ai in self.ai_agents])}")  # ✅ FIXED: .upper() not .UP()
        self.log_process("=" * 80)
        
        # Load parameters
        parameters = self.load_parameters(parameters_file)
        if not parameters:
            self.log_process("❌ No parameters loaded. Exiting.")
            return None
        
        self.log_process(f"📋 Loaded {len(parameters)} improvement parameters")
        parameter_groups = self.create_parameter_groups(parameters, 10)
        self.log_process(f"📦 Created {len(parameter_groups)} parameter groups")

        # 🆕 ADD THIS BLOCK:
        self.log_process("\n" + "=" * 80)
        self.log_process("🔍 DETECTING FEATURES FROM AIM + STEPS")
        self.log_process("=" * 80)
        self.detect_features_from_aim_steps(aim, initial_steps)
        time.sleep(1)
        
        # Analyze AIM-level competition
        self.log_process("\n" + "=" * 80)
        self.log_process("🏢 ANALYZING AIM-LEVEL COMPETITION")
        self.log_process("=" * 80)
        self.analyze_aim_competition(aim)
        time.sleep(2)
        
        # STEP 1: Generate enhanced plans for each AI
        self.log_process("\n" + "=" * 80)
        self.log_process("🔄 STEP 1: PLAN ENHANCEMENT BY ALL AIs")
        self.log_process("=" * 80)
        
        for ai_name in self.ai_agents:
            plan, score = self.step1_plan_generation(aim, initial_steps, parameters, ai_name)
            self.ai_plans[ai_name] = plan
            self.ai_scores[ai_name] = score
            time.sleep(2)  # Rate limiting
        
        # STEP 2: Improvement chain for each AI's enhanced plan
        self.log_process("\n" + "=" * 80)
        self.log_process("🔄 STEP 2: IMPROVEMENT CHAIN PROCESS")
        self.log_process("=" * 80)
        
        for ai_name in self.ai_agents:
            original_plan = self.ai_plans[ai_name]
            improvement_result = self.step2_improvement_chain(aim, initial_steps, original_plan, ai_name)
            self.improvement_chain_results[ai_name] = improvement_result
            time.sleep(2)  # Rate limiting

        # Collect all steps for summary analysis
        all_generated_steps = []
        for ai_name in self.ai_agents:
            if ai_name in self.improvement_chain_results:
                final_plan = self.improvement_chain_results[ai_name]["final_plan"]
                steps = self.parse_steps_simple(final_plan)
                all_generated_steps.extend(steps)
        
        # Analyze summary-level competition
        if all_generated_steps:
            self.log_process("\n" + "=" * 80)
            self.log_process("🏢 ANALYZING SUMMARY-LEVEL COMPETITION")
            self.log_process("=" * 80)
            self.analyze_summary_competition(all_generated_steps)
            time.sleep(2)
        
        # STEP 3: Select best plan
        self.log_process("\n" + "=" * 80)
        self.log_process("🔄 STEP 3: BEST ENHANCED PLAN SELECTION")
        self.log_process("=" * 80)
        
        final_result = self.step3_best_plan_selection()
        
        # Save results
        self.save_results_to_excel(final_result, aim, initial_steps)
        
        # Final summary
        execution_time = time.time() - start_time
        self.log_process(f"\n🎉 ENHANCEMENT PROCESS COMPLETED!")
        self.log_process(f"⏱️ Total execution time: {execution_time:.2f} seconds")
        self.log_process(f"🏆 Best Enhanced Plan: {final_result['best_ai'].upper() if final_result['best_ai'] else 'NONE'} ({final_result['best_score']}/100)")  # ✅ FIXED: .upper() not .UP()
        self.log_process(f"📄 Results saved to Excel file")
        
        return final_result

def process_plan_for_api(username, aim, steps, event_queue=None):
    """Process plan with real-time updates"""
    try:
        steps_list = [steps[key] for key in sorted(steps.keys())]
        
        # Create planner with event queue
        planner = EnhancedMultiAIPlanner(event_queue=event_queue)
        
        # Send initial event
        if event_queue:
            event_queue.put({
                'type': 'process_start',
                'data': {
                    'username': username,
                    'aim': aim,
                    'steps_count': len(steps_list),
                    'agents_count': len(planner.ai_agents)
                },
                'timestamp': datetime.now().isoformat()
            })
        
        # Load parameters
        parameters = planner.load_parameters(planner.parameters_file)
        if not parameters:
            if event_queue:
                event_queue.put({
                    'type': 'error',
                    'data': {'message': 'Failed to load parameters'},
                    'timestamp': datetime.now().isoformat()
                })
            return {"error": "Failed to load parameters file"}
        
        planner.aim = aim
        planner.initial_steps = steps_list
        planner.initialize_tracking_file(aim, steps_list)

        # 🆕 ADD THIS BLOCK:
        if event_queue:
            event_queue.put({
                'type': 'feature_detection_start',
                'data': {'phase': 'initial'},
                'timestamp': datetime.now().isoformat()
            })

        planner.detect_features_from_aim_steps(aim, steps_list)
        time.sleep(1)
        
        # STEP 1: Generate plans
        if event_queue:
            event_queue.put({
                'type': 'phase_start',
                'data': {
                    'phase': 1,
                    'description': 'Generating plans from all AI agents'
                },
                'timestamp': datetime.now().isoformat()
            })
        
        for ai_name in planner.ai_agents:
            plan, score = planner.step1_plan_generation(aim, steps_list, parameters, ai_name)
            planner.ai_plans[ai_name] = plan
            planner.ai_scores[ai_name] = score
            time.sleep(1)
        
        # STEP 2: Improvement chains
        if event_queue:
            event_queue.put({
                'type': 'phase_start',
                'data': {
                    'phase': 2,
                    'description': 'Running improvement chains'
                },
                'timestamp': datetime.now().isoformat()
            })
        
        for ai_name in planner.ai_agents:
            original_plan = planner.ai_plans[ai_name]
            improvement_result = planner.step2_improvement_chain(aim, steps_list, original_plan, ai_name)
            planner.improvement_chain_results[ai_name] = improvement_result
            time.sleep(1)
        
        # Format final response with competition data
        api_response = {
            'agents': {},
            'competition_analysis': {
                'aim_level': planner.aim_competition,
                'step_level': planner.mini_plan_competitions,
                'summary_level': planner.summary_competition
            }
        }
        
        for i, ai_name in enumerate(planner.ai_agents, 1):
            if ai_name in planner.improvement_chain_results:
                result = planner.improvement_chain_results[ai_name]
                final_plan = result["final_plan"]
                plan_steps = {}
                
                lines = final_plan.split('\n')
                step_count = 1
                for line in lines:
                    line = line.strip()
                    if line and (line.startswith('-') or line.startswith('•')):
                        step_text = line.lstrip('-•').strip()
                        step_text = re.sub(r'^step\s*\d+\s*:?\s*', '', step_text, flags=re.IGNORECASE)
                        if step_text:
                            plan_steps[f'step{step_count}'] = step_text
                            step_count += 1
                
                agent_data = {
                    "agent_name": ai_name,
                    "steps": plan_steps,
                    "original_score": planner.ai_scores.get(ai_name, 0),
                    "final_score": result.get("final_score", 0),
                    "average_score": round(result.get("average_score", 0), 2),
                    "all_improvement_scores": result.get("all_scores", []),
                    "total_improvements": len(result.get("all_scores", [])) - 1
                }
                api_response['agents'][f'agent{i}'] = agent_data
        
        # Send completion
        if event_queue:
            event_queue.put({
                'type': 'process_complete',
                'data': {
                    'total_agents': len(api_response['agents']),
                    'result': api_response
                },
                'timestamp': datetime.now().isoformat()
            })
        
        return api_response
        
    except Exception as e:
        if event_queue:
            event_queue.put({
                'type': 'error',
                'data': {'message': str(e)},
                'timestamp': datetime.now().isoformat()
            })
        return {"error": str(e)}

# Usage Example
if __name__ == "__main__":
    planner = EnhancedMultiAIPlanner()
    
    try:
        # Run the complete planning process with user input
        result = planner.run_complete_planning_process()
        
        if result is not None:
            print("\n" + "=" * 70)
            print("🎯 ORIGINAL STEPS:")
            print("=" * 30)
            for i, step in enumerate(planner.initial_steps, 1):
                print(f"{i}. {step}")
                
            print("\n" + "=" * 70)
            print("🏆 BEST ENHANCED PLAN:")
            print("=" * 70)
            print(result["best_plan"])
            
            print(f"\n📄 Complete process log saved to: {planner.tracking_file}")
            print(f"📊 Excel results saved to: ai_planning_results.xlsx")
            
            # Optional: Ask if user wants to see detailed scores
            show_details = input("\n📊 Show detailed scores? (y/n): ").strip().lower()
            if show_details == 'y':
                print("\n📊 DETAILED ENHANCEMENT SCORES:")
                print("-" * 50)
                for ai_name in planner.ai_agents:
                    if ai_name in planner.improvement_chain_results:
                        result_data = planner.improvement_chain_results[ai_name]
                        print(f"{ai_name.upper():<12} | Enhanced: {planner.ai_scores[ai_name]:>3}/100 | Final: {result_data['final_score']:>3}/100 | Average: {result_data['average_score']:>5.1f}/100")  # ✅ FIXED: .upper() not .UP()
                        
            # Ask if user wants to see step-by-step comparison
            show_comparison = input("\n🔄 Show before/after comparison? (y/n): ").strip().lower()
            if show_comparison == 'y':
                print("\n" + "=" * 70)
                print("📋 BEFORE vs AFTER COMPARISON:")
                print("=" * 70)
                print("🔴 ORIGINAL STEPS:")
                for i, step in enumerate(planner.initial_steps, 1):
                    print(f"   {i}. {step}")
                print(f"\n🟢 BEST ENHANCED PLAN ({result['best_ai'].upper()}):")  # ✅ FIXED: .upper() not .UP()
                print(result["best_plan"])
        else:
            print("❌ Enhancement process failed to complete properly.")
        
    except KeyboardInterrupt:
        print("\n⚠️ Process interrupted by user")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()






